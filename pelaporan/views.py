# pelaporan/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.gis.geos import Point, GEOSGeometry, GEOSException
from .models import LaporanJalan, FotoLaporan
from .forms import LaporanForm, FeedbackForm, RegisterForm, LoginForm
from django.contrib import messages
from pathlib import Path
from django.conf import settings
from django.core.mail import send_mail
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from datetime import datetime, timedelta

BASE_DIR = settings.BASE_DIR

# --- CACHING UNTUK BATAS BATAM ---
_batam_boundary_cache = None

def get_batam_boundary():
    """Fungsi helper untuk membaca WKT dari file, DENGAN CACHING."""
    global _batam_boundary_cache

    if _batam_boundary_cache is not None:
        print("Menggunakan Batas Batam dari cache.")
        return _batam_boundary_cache

    print("Membaca Batas Batam dari file...")
    WKT_FILE_PATH = BASE_DIR / 'batas_batam.wkt'
    boundary = None
    try:
        if WKT_FILE_PATH.is_file():
            with open(WKT_FILE_PATH, 'r', encoding='utf-8') as f:
                batas_wkt_string = f.read().strip()
            if batas_wkt_string:
                boundary = GEOSGeometry(batas_wkt_string, srid=4326)
                print("GEOSGeometry BERHASIL dibuat!")
            else:
                print("PERINGATAN (views-cache): File batas_batam.wkt kosong.")
        else:
            print(f"PERINGATAN (views-cache): File batas wilayah TIDAK DITEMUKAN di {WKT_FILE_PATH}")

    except GEOSException as e:
        print(f"ERROR GEOS (views-cache): Gagal membuat geometri dari WKT: {e}")
    except Exception as e:
        print(f"ERROR LAIN (views-cache) saat memuat batas Batam: {type(e).__name__} - {e}")

    _batam_boundary_cache = boundary
    return _batam_boundary_cache


# --- VIEWS APLIKASI ---

def halaman_peta_utama(request):
    """Menampilkan halaman peta utama (homepage)."""
    # ✅ TAMBAHAN: Hitung statistik
    total_laporan = LaporanJalan.objects.count()
    diverifikasi = LaporanJalan.objects.filter(status='DIVERIFIKASI').count()
    diperbaiki = LaporanJalan.objects.filter(status='DIPERBAIKI').count()
    selesai = LaporanJalan.objects.filter(status='SELESAI').count()
    
    context = {
        'title': 'Peta Laporan Jalan Rusak',
        'total_laporan': total_laporan,
        'diverifikasi': diverifikasi,
        'diperbaiki': diperbaiki,
        'selesai': selesai,
    }
    return render(request, 'pelaporan/peta_utama.html', context)


def data_laporan_geojson(request):
    """API endpoint untuk mengirim data GeoJSON ke Leaflet."""
    # ✅ TAMBAHAN: Filter berdasarkan query params
    status_filter = request.GET.get('status', None)
    jenis_filter = request.GET.get('jenis', None)
    tingkat_filter = request.GET.get('tingkat', None)
    
    laporan_query = LaporanJalan.objects.filter(
        status__in=['DIVERIFIKASI', 'DIPERBAIKI']
    )
    
    # Terapkan filter jika ada
    if status_filter:
        laporan_query = laporan_query.filter(status=status_filter)
    if jenis_filter:
        laporan_query = laporan_query.filter(jenis_kerusakan=jenis_filter)
    if tingkat_filter:
        laporan_query = laporan_query.filter(tingkat_kerusakan=tingkat_filter)
    
    features = []
    for laporan in laporan_query:
        foto_urls = [foto.foto.url for foto in laporan.foto_set.all()]
        features.append({
            "type": "Feature", 
            "id": laporan.id,
            "properties": {
                "deskripsi": laporan.deskripsi,
                "status": laporan.status,
                "jenis_kerusakan": laporan.get_jenis_kerusakan_display(),
                "tingkat_kerusakan": laporan.get_tingkat_kerusakan_display(),
                "tanggal_lapor": laporan.tanggal_lapor.strftime('%d/%m/%Y %H:%M'),
                "foto_urls": foto_urls
            },
            "geometry": {
                "type": "Point",
                "coordinates": [laporan.lokasi.x, laporan.lokasi.y]
            }
        })
    
    data_geojson = {"type": "FeatureCollection", "features": features}
    return JsonResponse(data_geojson)


def tambah_laporan(request):
    """Menampilkan halaman form (GET) atau memproses data POST via AJAX."""
    if request.method == 'POST':
        form = LaporanForm(request.POST, request.FILES)

        if form.is_valid():
            lat = form.cleaned_data['latitude']
            lon = form.cleaned_data['longitude']
            lokasi_point = Point(float(lon), float(lat), srid=4326)

            # Validasi Batas Wilayah
            batam_boundary_geom = get_batam_boundary()
            if batam_boundary_geom and not lokasi_point.within(batam_boundary_geom):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Lokasi laporan harus berada di dalam wilayah Batam.'
                }, status=400)
            elif not batam_boundary_geom:
                print("PERINGATAN (views): Tidak dapat memvalidasi batas Batam, laporan tetap disimpan.")

            # Simpan laporan
            laporan = form.save(commit=False)
            laporan.lokasi = lokasi_point
            
            # ✅ ASSIGN USER JIKA LOGIN
            if request.user.is_authenticated:
                laporan.user = request.user
                # Auto-fill email dari user jika tidak diisi
                if not laporan.email_pelapor:
                    laporan.email_pelapor = request.user.email
            
            laporan.save()

            # ✅ VALIDASI & SIMPAN MULTIPLE FOTO
            foto_list = request.FILES.getlist('foto_uploads')
            max_photos = 5  # Batasi maksimal 5 foto
            max_size = 5 * 1024 * 1024  # 5MB per foto
            
            foto_tersimpan = 0
            foto_error = []
            
            for idx, f in enumerate(foto_list[:max_photos]):  # Ambil max 5 foto pertama
                # Validasi ukuran
                if f.size > max_size:
                    foto_error.append(f"Foto {idx+1} terlalu besar (maks 5MB)")
                    continue
                
                # Validasi format
                allowed_formats = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
                if f.content_type not in allowed_formats:
                    foto_error.append(f"Foto {idx+1} format tidak didukung")
                    continue
                
                try:
                    FotoLaporan.objects.create(laporan=laporan, foto=f)
                    foto_tersimpan += 1
                except Exception as e:
                    foto_error.append(f"Gagal upload foto {idx+1}")
                    print(f"Error upload foto: {e}")
            
            # Pesan sukses
            success_message = f'Laporan Anda telah berhasil dikirim dan menunggu verifikasi (ID: #{laporan.id}).'
            if foto_tersimpan > 0:
                success_message += f' {foto_tersimpan} foto berhasil diupload.'
            if foto_error:
                success_message += f' Peringatan: {", ".join(foto_error)}'
            
            return JsonResponse({
                'status': 'success',
                'message': success_message,
                'laporan_id': laporan.id
            })

        else:
            print("Form Errors:", form.errors.as_json())
            errors_dict = {field: [e for e in errors] for field, errors in form.errors.items()}
            return JsonResponse({
                'status': 'form_error',
                'errors': errors_dict,
                'message': 'Harap perbaiki kesalahan pada form.'
            }, status=400)

    else:  # GET request
        # ✅ AUTO-FILL EMAIL JIKA USER LOGIN
        initial_data = {}
        if request.user.is_authenticated and request.user.email:
            initial_data['email_pelapor'] = request.user.email
        
        form = LaporanForm(initial=initial_data)
        context = {
            'form': form,
            'title': 'Lapor Jalan Rusak',
            'user_email': request.user.email if request.user.is_authenticated else None,  # ✅ Kirim email user
        }
        return render(request, 'pelaporan/form_laporan.html', context)


def faq_view(request):
    context = {'title': 'Pertanyaan Umum (FAQ)'}
    return render(request, 'pelaporan/faq.html', context)


def about_view(request):
    context = {'title': 'Tentang LaporJalan Batam'}
    return render(request, 'pelaporan/about.html', context)


def feedback_view(request):
    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            nama = form.cleaned_data['nama']
            email = form.cleaned_data['email']
            subjek = form.cleaned_data['subjek']
            pesan = form.cleaned_data['pesan']

            email_subject = f"Feedback LaporJalan: {subjek}"
            email_body = (
                f"Feedback baru:\n\n"
                f"Dari: {nama if nama else 'Anonim'}\n"
                f"Email: {email if email else '-'}\n\n"
                f"{pesan}"
            )
            
            try:
                send_mail(
                    email_subject,
                    email_body,
                    settings.EMAIL_HOST_USER,
                    [settings.EMAIL_HOST_USER],
                    fail_silently=False
                )
                messages.success(request, 'Terima kasih atas umpan balik Anda!')
                return redirect('halaman_peta_utama')  # ✅ FIXED: redirect sekarang ada
            except Exception as e:
                messages.error(request, f'Maaf, terjadi kesalahan saat mengirim pesan: {e}')
    else:
        form = FeedbackForm()
    
    context = {'form': form, 'title': 'Kirim Umpan Balik'}
    return render(request, 'pelaporan/feedback.html', context)


# ========== AUTH VIEWS ==========

def register_view(request):
    """Registrasi user baru"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Selamat datang, {user.first_name}! Akun Anda berhasil dibuat.')
            return redirect('dashboard')
    else:
        form = RegisterForm()
    
    context = {'form': form, 'title': 'Daftar Akun Baru'}
    return render(request, 'pelaporan/register.html', context)


def login_view(request):
    """Login user"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Selamat datang kembali, {user.first_name or user.username}!')
                
                # Redirect ke 'next' jika ada, atau ke dashboard
                next_url = request.GET.get('next', 'dashboard')
                return redirect(next_url)
        else:
            messages.error(request, 'Username atau password salah!')
    else:
        form = LoginForm()
    
    context = {'form': form, 'title': 'Login'}
    return render(request, 'pelaporan/login.html', context)


def logout_view(request):
    """Logout user"""
    logout(request)
    messages.info(request, 'Anda telah logout.')
    return redirect('halaman_peta_utama')


# ========== DASHBOARD VIEWS ==========

@login_required
def dashboard(request):
    """Dashboard user untuk melihat laporan mereka"""
    user = request.user
    
    # Ambil semua laporan user
    laporan_user = LaporanJalan.objects.filter(user=user).order_by('-tanggal_lapor')
    
    # Statistik
    total = laporan_user.count()
    baru = laporan_user.filter(status='BARU').count()
    diverifikasi = laporan_user.filter(status='DIVERIFIKASI').count()
    diperbaiki = laporan_user.filter(status='DIPERBAIKI').count()
    selesai = laporan_user.filter(status='SELESAI').count()
    
    context = {
        'title': 'Dashboard Saya',
        'laporan_list': laporan_user,
        'total': total,
        'baru': baru,
        'diverifikasi': diverifikasi,
        'diperbaiki': diperbaiki,
        'selesai': selesai,
    }
    return render(request, 'pelaporan/dashboard.html', context)


@login_required
def detail_laporan(request, laporan_id):
    """Detail laporan milik user"""
    laporan = get_object_or_404(LaporanJalan, id=laporan_id, user=request.user)
    
    context = {
        'title': f'Detail Laporan #{laporan.id}',
        'laporan': laporan,
    }
    return render(request, 'pelaporan/detail_laporan.html', context)


@login_required
def edit_laporan(request, laporan_id):
    """Edit laporan (hanya jika status masih BARU)"""
    laporan = get_object_or_404(LaporanJalan, id=laporan_id, user=request.user)
    
    # Cek status - hanya BARU yang bisa diedit
    if laporan.status != 'BARU':
        messages.error(request, 'Laporan yang sudah diverifikasi tidak dapat diedit!')
        return redirect('detail_laporan', laporan_id=laporan_id)
    
    if request.method == 'POST':
        form = LaporanForm(request.POST, request.FILES, instance=laporan)
        if form.is_valid():
            updated_laporan = form.save(commit=False)
            
            # Update lokasi jika ada perubahan koordinat
            if 'latitude' in form.cleaned_data and 'longitude' in form.cleaned_data:
                lat = form.cleaned_data['latitude']
                lon = form.cleaned_data['longitude']
                updated_laporan.lokasi = Point(float(lon), float(lat), srid=4326)
            
            updated_laporan.save()
            
            # Handle foto baru jika ada
            foto_list = request.FILES.getlist('foto_uploads')
            for f in foto_list[:5]:
                FotoLaporan.objects.create(laporan=updated_laporan, foto=f)
            
            messages.success(request, 'Laporan berhasil diperbarui!')
            return redirect('detail_laporan', laporan_id=laporan_id)
    else:
        # Pre-fill koordinat dari lokasi existing
        initial_data = {
            'latitude': laporan.lokasi.y if laporan.lokasi else None,
            'longitude': laporan.lokasi.x if laporan.lokasi else None,
        }
        form = LaporanForm(instance=laporan, initial=initial_data)
    
    context = {
        'title': f'Edit Laporan #{laporan.id}',
        'form': form,
        'laporan': laporan,
        'is_edit': True,
    }
    return render(request, 'pelaporan/form_laporan.html', context)


@login_required
def hapus_laporan(request, laporan_id):
    """Hapus laporan (hanya jika status masih BARU)"""
    laporan = get_object_or_404(LaporanJalan, id=laporan_id, user=request.user)
    
    # Cek status - hanya BARU yang bisa dihapus
    if laporan.status != 'BARU':
        messages.error(request, 'Laporan yang sudah diverifikasi tidak dapat dihapus!')
        return redirect('detail_laporan', laporan_id=laporan_id)
    
    if request.method == 'POST':
        laporan.delete()
        messages.success(request, 'Laporan berhasil dihapus!')
        return redirect('dashboard')
    
    context = {
        'title': 'Konfirmasi Hapus Laporan',
        'laporan': laporan,
    }
    return render(request, 'pelaporan/hapus_laporan.html', context)


@login_required
def hapus_foto_laporan(request, foto_id):
    """Hapus foto dari laporan (AJAX)"""
    if request.method == 'POST':
        foto = get_object_or_404(FotoLaporan, id=foto_id, laporan__user=request.user)
        
        # Cek status laporan
        if foto.laporan.status != 'BARU':
            return JsonResponse({
                'status': 'error',
                'message': 'Tidak dapat menghapus foto dari laporan yang sudah diverifikasi!'
            }, status=400)
        
        foto.delete()
        return JsonResponse({
            'status': 'success',
            'message': 'Foto berhasil dihapus!'
        })
    
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)


# ========== ADMIN STATISTICS & ANALYTICS ==========

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.db.models import Count, Q, Avg
from django.db.models.functions import TruncMonth, TruncDate
import json
from datetime import timedelta

@staff_member_required
def admin_statistics(request):
    """Dashboard statistik untuk admin/staff"""
    
    # Total statistik
    total_laporan = LaporanJalan.objects.count()
    baru = LaporanJalan.objects.filter(status='BARU').count()
    diverifikasi = LaporanJalan.objects.filter(status='DIVERIFIKASI').count()
    diperbaiki = LaporanJalan.objects.filter(status='DIPERBAIKI').count()
    selesai = LaporanJalan.objects.filter(status='SELESAI').count()
    
    # Laporan 7 hari terakhir
    week_ago = datetime.now() - timedelta(days=7)
    laporan_minggu_ini = LaporanJalan.objects.filter(tanggal_lapor__gte=week_ago).count()
    
    # Laporan per jenis kerusakan
    per_jenis = LaporanJalan.objects.values('jenis_kerusakan').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Laporan per tingkat kerusakan
    per_tingkat = LaporanJalan.objects.values('tingkat_kerusakan').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Laporan per bulan (6 bulan terakhir)
    six_months_ago = datetime.now() - timedelta(days=180)
    per_bulan = LaporanJalan.objects.filter(
        tanggal_lapor__gte=six_months_ago
    ).annotate(
        bulan=TruncMonth('tanggal_lapor')
    ).values('bulan').annotate(
        total=Count('id')
    ).order_by('bulan')
    
    # Konversi ke format Chart.js
    labels_bulan = [item['bulan'].strftime('%b %Y') for item in per_bulan]
    data_bulan = [item['total'] for item in per_bulan]
    
    # User paling aktif (top 5)
    top_users = LaporanJalan.objects.filter(user__isnull=False).values(
        'user__username', 'user__first_name', 'user__email'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Rata-rata waktu penanganan (dari DIVERIFIKASI ke SELESAI)
    # Ini butuh tracking tanggal per status, untuk sekarang kita skip dulu
    
    # Laporan dengan foto vs tanpa foto
    dengan_foto = LaporanJalan.objects.filter(foto_set__isnull=False).distinct().count()
    tanpa_foto = total_laporan - dengan_foto
    
    context = {
        'title': 'Dashboard Admin - Statistik Laporan',
        'total_laporan': total_laporan,
        'baru': baru,
        'diverifikasi': diverifikasi,
        'diperbaiki': diperbaiki,
        'selesai': selesai,
        'laporan_minggu_ini': laporan_minggu_ini,
        'per_jenis': per_jenis,
        'per_tingkat': per_tingkat,
        'labels_bulan': json.dumps(labels_bulan),
        'data_bulan': json.dumps(data_bulan),
        'top_users': top_users,
        'dengan_foto': dengan_foto,
        'tanpa_foto': tanpa_foto,
    }
    return render(request, 'pelaporan/admin_statistics.html', context)


@staff_member_required
def export_laporan_excel(request):
    """Export semua laporan ke Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Buat workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Jalan"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'ID', 'Tanggal Lapor', 'Status', 'Jenis Kerusakan', 'Tingkat', 
        'Deskripsi', 'Latitude', 'Longitude', 'Email Pelapor', 
        'User', 'Jumlah Foto', 'Terakhir Update'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    laporan_list = LaporanJalan.objects.all().select_related('user').prefetch_related('foto_set')
    
    for row_num, laporan in enumerate(laporan_list, 2):
        ws.cell(row=row_num, column=1, value=laporan.id)
        ws.cell(row=row_num, column=2, value=laporan.tanggal_lapor.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=3, value=laporan.get_status_display())
        ws.cell(row=row_num, column=4, value=laporan.get_jenis_kerusakan_display())
        ws.cell(row=row_num, column=5, value=laporan.get_tingkat_kerusakan_display())
        ws.cell(row=row_num, column=6, value=laporan.deskripsi or '-')
        ws.cell(row=row_num, column=7, value=f"{laporan.lokasi.y:.6f}" if laporan.lokasi else '-')
        ws.cell(row=row_num, column=8, value=f"{laporan.lokasi.x:.6f}" if laporan.lokasi else '-')
        ws.cell(row=row_num, column=9, value=laporan.email_pelapor or '-')
        ws.cell(row=row_num, column=10, value=laporan.user.username if laporan.user else 'Guest')
        ws.cell(row=row_num, column=11, value=laporan.foto_set.count())
        ws.cell(row=row_num, column=12, value=laporan.tanggal_update.strftime('%d/%m/%Y %H:%M'))
    
    # Auto-adjust column width
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=laporan_jalan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@staff_member_required
def export_laporan_csv(request):
    """Export laporan ke CSV"""
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=laporan_jalan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'ID', 'Tanggal Lapor', 'Status', 'Jenis Kerusakan', 'Tingkat', 
        'Deskripsi', 'Latitude', 'Longitude', 'Email Pelapor', 
        'User', 'Jumlah Foto', 'Terakhir Update'
    ])
    
    # Data
    laporan_list = LaporanJalan.objects.all().select_related('user').prefetch_related('foto_set')
    
    for laporan in laporan_list:
        writer.writerow([
            laporan.id,
            laporan.tanggal_lapor.strftime('%d/%m/%Y %H:%M'),
            laporan.get_status_display(),
            laporan.get_jenis_kerusakan_display(),
            laporan.get_tingkat_kerusakan_display(),
            laporan.deskripsi or '-',
            f"{laporan.lokasi.y:.6f}" if laporan.lokasi else '-',
            f"{laporan.lokasi.x:.6f}" if laporan.lokasi else '-',
            laporan.email_pelapor or '-',
            laporan.user.username if laporan.user else 'Guest',
            laporan.foto_set.count(),
            laporan.tanggal_update.strftime('%d/%m/%Y %H:%M'),
        ])
    
    return response


@staff_member_required
def heatmap_data(request):
    """API untuk data heatmap (intensitas kerusakan per area)"""
    
    # Ambil semua laporan dengan koordinat
    laporan_list = LaporanJalan.objects.exclude(
        status='SELESAI'
    ).values('lokasi', 'tingkat_kerusakan')
    
    # Convert ke format heatmap
    heatmap_points = []
    for laporan in laporan_list:
        if laporan['lokasi']:
            # Weight berdasarkan tingkat kerusakan
            weight = 1
            if laporan['tingkat_kerusakan'] == 'SEDANG':
                weight = 2
            elif laporan['tingkat_kerusakan'] == 'BERAT':
                weight = 3
            
            # Extract lat/lon dari POINT
            from django.contrib.gis.geos import GEOSGeometry
            point = GEOSGeometry(laporan['lokasi'])
            
            heatmap_points.append({
                'lat': point.y,
                'lng': point.x,
                'weight': weight
            })
    
    return JsonResponse({'points': heatmap_points})


@staff_member_required
def laporan_list_admin(request):
    """List semua laporan untuk admin dengan filter"""
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    jenis_filter = request.GET.get('jenis', '')
    tingkat_filter = request.GET.get('tingkat', '')
    search = request.GET.get('search', '')
    
    # Base query
    laporan_list = LaporanJalan.objects.all().select_related('user').prefetch_related('foto_set')
    
    # Apply filters
    if status_filter:
        laporan_list = laporan_list.filter(status=status_filter)
    if jenis_filter:
        laporan_list = laporan_list.filter(jenis_kerusakan=jenis_filter)
    if tingkat_filter:
        laporan_list = laporan_list.filter(tingkat_kerusakan=tingkat_filter)
    if search:
        laporan_list = laporan_list.filter(
            Q(id__icontains=search) |
            Q(deskripsi__icontains=search) |
            Q(email_pelapor__icontains=search) |
            Q(user__username__icontains=search)
        )
    
    # Order by newest
    laporan_list = laporan_list.order_by('-tanggal_lapor')
    
    context = {
        'title': 'Kelola Semua Laporan',
        'laporan_list': laporan_list,
        'status_choices': LaporanJalan.STATUS_CHOICES,
        'jenis_choices': LaporanJalan.JENIS_CHOICES,
        'tingkat_choices': LaporanJalan.TINGKAT_CHOICES,
        'current_filters': {
            'status': status_filter,
            'jenis': jenis_filter,
            'tingkat': tingkat_filter,
            'search': search,
        }
    }
    return render(request, 'pelaporan/admin_laporan_list.html', context)